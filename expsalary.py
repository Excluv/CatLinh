import sqlite3
import openpyxl
import pandas as pd
from copy import copy
from datetime import datetime

def copysheet(source_sheet, target_sheet):
    copysheet_attributes(source_sheet, target_sheet)
    copycells(source_sheet, target_sheet)


def copysheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # Set row dimensions:
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column width')
    else:
        target_sheet.sheet_format.defaultColWidth = \
            copy(source_sheet.sheet_format.defaultColWidth)

    # Set specific column width and hidden property:
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = \
            copy(source_sheet.column_dimensions[key].min)
        target_sheet.column_dimensions[key].max = \
            copy(source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = \
            copy(source_sheet.column_dimensions[key].width)
        target_sheet.column_dimensions[key].hidden = \
            copy(source_sheet.column_dimensions[key].hidden)


def copycells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(row=row, column=col)
        target_cell.value = source_cell.value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell.hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


conn = sqlite3.connect('CatLinh.sqlite')
cur = conn.cursor()

# from_date = input('From: ')
# to_date = input('To: ')

from_date = input('From date: ')
to_date = input('To date: ')

cur.execute(f'''
SELECT Orders.ID, Orders.Container_ID, Destination.Name, Type.Name,
        PickupLocation.Name, DropoffLocation.Name, Orders.From_Date,
        Orders.To_Date, Orders.Provision, Salary.Salary, PickupLocation.Cost,
        DropoffLocation.Cost
FROM Orders JOIN Destination JOIN Type JOIN PickupLocation JOIN
        DropoffLocation JOIN Salary
WHERE Orders.From_Date BETWEEN '{from_date}' AND '{to_date}' AND
        Orders.Vehicle = Salary.Vehicle AND
        Orders.Destination_ID = Destination.ID AND
        Type.ID = Salary.Type_ID AND
        Destination.ID = Salary.Destination_ID AND
        Orders.PUL_ID = PickupLocation.ID AND
        Orders.DOL_ID = DropoffLocation.ID
''')

data = cur.fetchall()
df = pd.DataFrame(data, columns=
                        ['STT', 'Số Container', 'Tuyến VC',
                         'Loại Cont', 'Cảng nâng', 'Cảng hạ',
                         'Ngày đóng/rút hàng', 'Giải phóng xe',
                         'Tạm ứng sản xuất', 'Cước VC',
                         'Phí nâng', 'Phí hạ'])

df.to_excel('C:\\Users\\jio\\Desktop\\Test3.xlsx',
            sheet_name='51C-774.99',index=False, startrow=4)

title = f'BẢNG LƯƠNG THÁNG {int(datetime.now().month) - 1}/{datetime.now().year}'
info = 'Số xe: 51C-774.99                 Họ & tên tài xế: Hà Thanh Phước'

target_file = openpyxl.load_workbook('C:\\Users\\jio\\Desktop\\Test3.xlsx')
xcel = target_file.active

xcel.cell(row=1, column=1).value = title
xcel.merge_cells('A1:Q1')
xcel.cell(row=3, column=1).value = info

additional_cols = ['Phí phát sinh VC', 'Ghi chú',
                  'Tổng cộng', 'Số lượng dầu cấp', 'Ngày']

col_index = 13
for x in range(len(additional_cols)):
    xcel.cell(row=5, column=col_index + x).value = additional_cols[x]

index = 6
for row in range(len(data)):
    sum = 0
    for col in range(10, 13):
        try:
            sum += xcel.cell(row=index + row, column=col).value
        except:
            pass
    xcel.cell(row=index + row, column=15).value = sum

footer_index = index + len(data)
xcel.cell(row=footer_index, column=1).value='Tổng cộng'
xcel.merge_cells(f'A{footer_index}:H{footer_index}')
for col in range(9, 16):
    sum = 0
    for row in range(6, footer_index):
        if xcel.cell(row=row, column=col).value is None: continue
        sum += xcel.cell(row=row, column=col).value
    xcel.cell(row=footer_index, column=col).value = sum

footer_index = 31
source_file= openpyxl.load_workbook('C:\\Users\\jio\\Desktop\\CatLinhProject\\Samples\\Bảng lương T04.xlsx')
source_sheet = source_file['51C-774.99']
for row in range(footer_index, footer_index + 21):
    for col in range(1, 6):
        xcel.cell(row=row, column=col).value = \
            source_sheet.cell(row=row, column=col).value

source_sheet = source_file['Tổng cộng']
target_file.create_sheet('Tổng cộng')
copysheet(source_file['Tổng cộng'], target_file['Tổng cộng'])

target_file.save('C:\\Users\\jio\\Desktop\\Test3.xlsx')
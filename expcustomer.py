import openpyxl
import pandas as pd
import sqlite3
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment


def data_to_excel():
    # Export main data to an excel file:
    conn = sqlite3.connect('file:CatLinh.sqlite?mode=ro', uri=True)
    cur = conn.cursor()

    cur.execute('''
    SELECT Orders.ID, Orders.Container_ID, Destination.Name, PickupLocation.Name,
           DropoffLocation.Name, Orders.From_Date, Orders.To_Date,
           Revenue.Revenue, PickupLocation.Cost, DropoffLocation.Cost,
           Orders.Incurred_Cost, Orders.Description
    FROM Orders JOIN Destination JOIN PickupLocation
                JOIN DropoffLocation JOIN Revenue
    WHERE Orders.Destination_ID = Destination.ID AND Orders.PUL_ID = PickupLocation.ID AND
          Orders.DOL_ID = DropoffLocation.ID AND Orders.Type_ID = Revenue.Type_ID AND
          Orders.Destination_ID = Revenue.Destination_ID AND Orders.Customer_ID = Revenue.Customer_ID
    ''')

    data = cur.fetchall()

    df = pd.DataFrame(data, columns=['STT', 'Số Container', 'Kho Hàng', 'Cảng nâng', 'Cảng hạ',
                                     'Đóng/rút hàng', 'Giải phóng xe', 'Cước VC', 'Phí nâng', 'Phí hạ',
                                     'Phí phát sinh', 'Ghi chú'])

    df.to_excel('C:\\Users\\jio\\Desktop\\Test.xlsx', index=False, startrow=8)

    # Interact with other parts of the file:
    file = openpyxl.load_workbook('C:\\Users\\jio\\Desktop\\Test.xlsx')
    xcel = file.active

    # Input headers, title, receiver:
    headers = ['CÔNG TY TNHH THƯƠNG MẠI DỊCH VẬN TẢI CÁT LINH',
               'MST: 0306980684',
               'Địa chỉ: 441/112 Điện Biên Phủ, P.25, Q.Bình Thạnh, TP.HCM']
    title = f'BẢNG THỐNG KÊ VẬN CHUYỂN THÁNG 0{int(datetime.now().month) - 1}/{datetime.now().year}'
    receiver = 'Kính gửi: CÔNG TY TNHH LIÊN ĐẠI PHÁT'

    for row in range(8):
        if row < 3:
            xcel.cell(row=row + 1, column=1).value = headers[row]
        if (row == 3) or (row == 5):
            continue
        if row == 4:
            xcel.cell(row=row + 1, column=1).value = title
            xcel.cell(row=row + 1, column=1).alignment = Alignment(horizontal='center')
            xcel.merge_cells('A5:M5')
            continue
        if row == 6:
            xcel.cell(row=row + 1, column=1).value = receiver
            continue

    # Deal with body details:
    phrase = 'Tổng cộng'
    index = 10 + len(data)
    xcel.cell(row=index, column=1).value = phrase
    xcel.merge_cells(f'A{index}:G{index}')

    total_sum = 0
    for col in range(8, 12):
        sum = 0
        for row in range(10, index):
            if xcel.cell(row=row, column=col).value is None:
                continue

            sum += xcel.cell(row=row, column=col).value

        xcel.cell(row=index, column=col).value = sum
        total_sum += sum

    xcel['M9'] = 'Tổng cộng'
    xcel[f'M{index}'] = total_sum
    for row in range(10, index):
        sum = 0
        for col in range(8, 12):
            if xcel.cell(row=row, column=col).value is None: continue
            sum += xcel.cell(row=row, column=col).value

        xcel.cell(row=row, column=13).value = sum

    # Set footer:
    footer_index = index + 2
    footers = ['* Cước vận chuyển:', '* Phí nâng:', '* Phí hạ:', '* Phí phát sinh:', '* Tổng cộng']

    for num in range(5):
        xcel.cell(row=footer_index + num, column=1).value = footers[num]
        xcel.cell(row=footer_index + num, column=5).value = xcel.cell(row=index, column=8 + num).value
        if num == 4:
            xcel.cell(row=footer_index + num, column=5).value = xcel.cell(row=index, column=13).value

    debt_index = footer_index + 6
    debts = [f'* Nợ cuối T0{int(datetime.now().month) - 2}/2022:',
             f'* Phát sinh T0{int(datetime.now().month) - 1}/2022:',
             f'* Thanh toán T0{int(datetime.now().month) - 1}/2022:',
             f'* Nợ cuối T0{int(datetime.now().month) - 1}/2022:']

    for num in range(4):
        xcel.cell(row=debt_index + num, column=1).value = debts[num]

    file.save('C:\\Users\\jio\\Desktop\\Test.xlsx')


data_to_excel()